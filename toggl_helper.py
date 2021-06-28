"""Toggl Helper"""
import datetime
import json
import logging
import os
import sys
from base64 import b64encode
from copy import deepcopy
from getpass import getuser

import requests
import rumps
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from requests import HTTPError

BASE_URL = 'https://api.track.toggl.com/api/v8/'
FULL_USER_DATA_ENDPOINT = BASE_URL + 'me?with_related_data=true'
TIME_ENTRIES_ENDPOINT = BASE_URL + 'time_entries'
TOGGL_API_KEY_PATH = f'/Users/{getuser()}/.thk.txt'
WORKBOOK_PATH_BASE = f'/Users/{getuser()}/Desktop/toggl_workbook_DATE.xlsx'
WORKSHEET_HEADERS = ['TASK', 'PROJECT', 'START', 'END', 'DURATION (optional)']
EXAMPLE_TASK = 'Example task (no need to delete)'
EXAMPLE_PROJECT = 'Example project (no need to delete)'
WORKSHEET_ROW_EXAMPLE = [EXAMPLE_TASK, EXAMPLE_PROJECT,
                         datetime.datetime.now().replace(microsecond=0),
                         datetime.datetime.now().replace(microsecond=0),
                         '=D2-C2']


def get_workbook_path():
    return WORKBOOK_PATH_BASE.replace('DATE', str(datetime.date.today()))


class TogglHelper(object):

    projects = {0: {'name': None}, 1: {'name': None}}
    # the popup for the project selection have the buttons with ID 0 and 1 connected to no project

    def __init__(self):
        self.config = {
            "app_name": "Toggl Helper",
            "app_icon": "toggl_icon.png",
            "start_track": "Start Tracking",
            "start_new": "Start New Tracking",
            "stop_track": "Stop Tracking",
            "create_empty_workbook": "Create Empty Workbook",
            "available_projects": "Show Available Projects",
            "sync": "Sync Toggl",
            "interval": 60
        }
        # self.app = rumps.App(self.config["app_name"], icon=self.config["app_icon"])
        self.app = rumps.App(self.config["app_name"])
        self.timer = rumps.Timer(self.on_tick, 60)
        self.interval = self.config["interval"]
        self.app.title = ""
        self.start_button = rumps.MenuItem(title=self.config["start_track"], callback=self.start_timer)
        self.start_new_button = rumps.MenuItem(title=self.config["start_new"], callback=None)
        self.stop_button = rumps.MenuItem(title=self.config["stop_track"], callback=None)
        self.create_button = rumps.MenuItem(title=self.config["create_empty_workbook"], callback=self.get_workbook)
        self.available_projects = rumps.MenuItem(
            title=self.config["available_projects"], callback=self.show_toggle_projects_names)
        self.sync_button = rumps.MenuItem(title=self.config["sync"], callback=self.sync_toggl)
        self.app.menu = [self.start_button, self.start_new_button, self.stop_button,
                         self.create_button, self.available_projects, self.sync_button]
        self.current_task = {header: ... for header in WORKSHEET_HEADERS}
        self.task_to_record = None
        self.toggl_headers = self.get_toggl_headers()
        self.projects.update({key: project for key, project in enumerate(self.get_toggl_projects(), start=2)})

    def get_toggl_headers(self):
        try:
            with open(TOGGL_API_KEY_PATH) as f:
                toggl_api_key = f.read()
        except FileNotFoundError:
            with open(TOGGL_API_KEY_PATH, 'w') as f:
                window = rumps.Window(dimensions=(450, 160))
                window.title = 'TOGGL API KEY'
                window.message = 'Please add your Toggle API key below.'
                window.default_text = ''
                response = window.run()
                toggl_api_key = response.text
                f.write(toggl_api_key)
        toggl_token = "Basic " + b64encode(f'{toggl_api_key}:api_token'.encode()).decode('ascii').rstrip()
        toggl_headers = {"Content-Type": "application/json", "Authorization": toggl_token}
        return toggl_headers

    def on_tick(self, sender):
        self.app.title = self.app.title[:-4] + str(datetime.timedelta(seconds=sender.elapsed_time))[:-3]
        sender.elapsed_time += 60

    def start_timer(self, sender):
        try:
            if sender.title == self.config["start_new"]:
                self.task_to_record = deepcopy(self.current_task)
            if self.input_project_details():
                if sender.title == self.config["start_new"]:
                    self.timer.stop()
                    self.record_task()
                self.timer.elapsed_time = 0
                self.start_button.set_callback(None)
                self.start_new_button.set_callback(self.start_timer)
                self.stop_button.set_callback(self.stop_timer)
                self.timer.start()
                self.current_task.update({'START': datetime.datetime.now().replace(microsecond=0)})
        except:
            logging.exception(datetime.datetime.now().replace(microsecond=0))

    def record_task(self):
        workbook = self.get_workbook()
        worksheet = workbook.active
        self.task_to_record.update({'END': datetime.datetime.now().replace(microsecond=0)})
        # duration = self.current_task['END'] - self.current_task['START']
        last_row = str(worksheet.max_row + 1)
        duration = f'=D{last_row}-C{last_row}'
        self.task_to_record.update({'DURATION (optional)': duration})
        worksheet.append(list(self.task_to_record.values()))
        worksheet[f'E{last_row}'].number_format = 'H:MM:SS'
        workbook.save(get_workbook_path())

    def get_workbook(self, sender=''):
        try:
            if sender.title == self.config['create_empty_workbook']:
                if os.path.exists(get_workbook_path()):
                    rumps.notification(
                        title="😳😳 Oops! File already exist.", subtitle="Please, check your desktop.", message="")
                else:
                    self.create_workbook()
                    rumps.notification(
                        title="🎉🎉 Great! File has been created.", subtitle="Please, check your desktop.", message="")
            else:
                return load_workbook(get_workbook_path())
        except FileNotFoundError:
            return self.create_workbook()
        except:
            logging.exception(datetime.datetime.now().replace(microsecond=0))

    def create_workbook(self):
        # workbook = Workbook()
        # worksheet = workbook.active
        # worksheet.append(WORKSHEET_HEADERS)
        # worksheet.append(WORKSHEET_ROW_EXAMPLE)
        # worksheet['E2'].number_format = 'H:MM:SS'
        # for col in worksheet.iter_cols():
        #     worksheet.column_dimensions[col[0].column_letter].width = 40
        # worksheet.auto_filter.ref = "A1:E1"
        # workbook.save(get_workbook_path())
        # return workbook
        workbook = Workbook()
        # SHEET 1
        worksheet = workbook.active
        worksheet.append(WORKSHEET_HEADERS)
        worksheet.append(WORKSHEET_ROW_EXAMPLE)
        worksheet['E2'].number_format = 'H:MM:SS'
        for col in worksheet.iter_cols():
            worksheet.column_dimensions[col[0].column_letter].width = 40
        worksheet.auto_filter.ref = "A1:E1"
        # SHEET 2
        worksheet_two = workbook.create_sheet()
        for project in self.projects.values():
            if project['name']:
                worksheet_two.append([project['name']])
        for col in worksheet.iter_cols():
            worksheet_two.column_dimensions[col[0].column_letter].width = 70
        # ADD DATA VALIDATION
        data_validation = DataValidation(type="list", formula1='=Sheet1!$A$1:$A$20')
        worksheet.add_data_validation(data_validation)
        data_validation.add('B3:B100')
        workbook.save(get_workbook_path())
        return workbook

    def input_project_details(self):
        ok_button_text = 'Please select the relevant project below.'
        window = rumps.Window(ok=ok_button_text, cancel=True, dimensions=(450, 160))
        # window.icon = self.config["app_icon"]
        window.title = 'TOGGL INPUT'
        window.message = 'Please add a task description below.'
        window.default_text = ''
        for project in self.projects.values():
            if project['name']:
                window.add_button(project['name'])
        response = window.run()
        input_project = self.projects[response.clicked]['name']
        input_task = response.text
        if input_project:
            self.current_task.update({'TASK': input_task, 'PROJECT': input_project})
            newline = '\n'
            self.app.title = f' {input_task.strip(newline)} | 0:00'
        return input_project

    def stop_timer(self, sender):
        try:
            self.timer.stop()
            self.task_to_record = deepcopy(self.current_task)
            self.record_task()
            self.start_button.set_callback(self.start_timer)
            self.start_new_button.set_callback(None)
            self.stop_button.set_callback(None)
            self.app.title = ""
        except:
            logging.exception(datetime.datetime.now().replace(microsecond=0))

    def sync_toggl(self, sender):
        try:
            workbook = load_workbook(get_workbook_path())
            worksheet = workbook.active
            for row in worksheet.iter_rows(min_row=2):
                # raw_duration = row[4].value.seconds
                description = row[0].value
                if description == EXAMPLE_TASK:
                    continue
                project = row[1].value
                start_hour, start_minute = round_time(row[2].value)
                stop_hour, stop_minute = round_time(row[3].value)
                start = datetime.datetime(
                    year=row[2].value.year,
                    month=row[2].value.month,
                    day=row[2].value.day,
                    hour=start_hour - 2,
                    minute=start_minute
                )
                stop = datetime.datetime(
                    year=row[3].value.year,
                    month=row[3].value.month,
                    day=row[3].value.day,
                    hour=stop_hour - 2,
                    minute=stop_minute
                )
                duration = stop - start
                project_key = self.get_project_dictionary_key(project)
                data = {
                    "time_entry": {
                        "description": description,
                        "duration": duration.seconds,
                        "start": start.isoformat() + '.000Z',
                        "stop": stop.isoformat() + '.000Z',
                        "pid": self.projects[project_key]['id'],
                        "created_with": "toggl-helper",
                        "billable": self.projects[project_key]['billable']
                    }
                }
                if duration:  # if duration more than 0:00 upload data
                    binary_data = json.JSONEncoder().encode(data).encode('utf-8')
                    requests.post(url=TIME_ENTRIES_ENDPOINT, data=binary_data, headers=self.toggl_headers)
            rumps.notification(title='🎉🎉 Great! Toggl has been synced.', subtitle='', message='')
        except:
            logging.exception(datetime.datetime.now().replace(microsecond=0))

    def get_project_dictionary_key(self, project_name: str):
        """Get key for project in self.projects"""
        for key, project in self.projects.items():
            if project['name'] == project_name:
                return key

    def show_toggle_projects_names(self, sender):
        """Print available projects in notification."""
        try:
            projects = '\n'.join([project['name'] for project in self.get_toggl_projects()])
            window = rumps.Window(message=projects, dimensions=(450, 0))
            # window.icon = self.config["app_icon"]
            window.title = 'TOGGL PROJECTS'
            window.run()
        except:
            logging.exception(datetime.datetime.now().replace(microsecond=0))

    def get_toggl_projects(self):
        """Get available Toggl projects."""
        response = requests.get(url=FULL_USER_DATA_ENDPOINT, headers=self.toggl_headers)
        try:
            response.raise_for_status()
        except HTTPError as e:
            os.remove(TOGGL_API_KEY_PATH)
            rumps.notification(title='You have entered a wrong API key.', subtitle='Please open the program again',
                               message='and add again your API key')
            sys.exit()
        return response.json()['data']['projects']

    def run(self):
        self.app.run(debug='debug')
        # self.app.run()


def round_time(time):
    """Round up or down time based on set break_point."""
    minutes = time.minute if time.second < 30 else time.minute + 1
    hours = time.hour
    break_point = 7  # minutes
    if minutes > 45 + break_point:
        hours += 1
        minutes = 0
    elif minutes > 30 + break_point:
        minutes = 45
    elif minutes > 15 + break_point:
        minutes = 30
    elif minutes > 0 + break_point:
        minutes = 15
    else:
        minutes = 0
    return hours, minutes


if __name__ == '__main__':
    logging.basicConfig(filename=f'/Users/{getuser()}/.th.log')
    toggl_helper = TogglHelper()
    toggl_helper.run()

